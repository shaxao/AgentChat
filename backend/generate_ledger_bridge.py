#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Python 桥接脚本：用 openpyxl 生成台账
替代 Java POI，避免 LOCAL_YEAR_FORMAT 等命名范围错误

核心策略：
  1. 用 data_only=True 打开模板读取缓存值（用于匹配编码）
  2. 用普通模式打开模板写入数据（保留公式结构）
  3. 通过行号关联两个 workbook

用法：
    python generate_ledger_bridge.py <template_path> <output_path> <rows_json_path>
"""

import sys
import json
import datetime
import openpyxl
from openpyxl import load_workbook


def normalize_code(s):
    """规范化编码：去.0后缀、提取纯数字、去前导零"""
    if not s:
        return ''
    s = str(s).strip()
    while s.endswith('.0'):
        s = s[:-2]
    digits = ''.join(c for c in s if c.isdigit())
    if not digits:
        return s
    stripped = digits.lstrip('0')
    return stripped if stripped else '0'


def is_code_like(s):
    """判断字符串是否像商品编码"""
    if not s or not s.strip():
        return False
    t = s.strip()
    # 如果是公式（以=开头），不是编码值
    if t.startswith('='):
        return False
    digits = ''.join(c for c in t if c.isdigit())
    others = ''.join(c for c in t if c not in '0123456789 .-/')
    if others:
        return False
    return 5 <= len(digits) <= 8


def find_header_row(sheet, max_scan=15):
    """扫描前N行，找到包含编码/代码/名称等关键词的表头行"""
    for r in range(1, min(max_scan, sheet.max_row) + 1):
        match_count = 0
        for c in range(1, min(sheet.max_column + 1, 30)):
            val = str(sheet.cell(row=r, column=c).value or '').strip()
            if any(kw in val for kw in ['编码', '代码', '货号', '名称', '品名']):
                match_count += 1
        if match_count >= 2:
            return r
    for r in range(1, min(max_scan, sheet.max_row) + 1):
        for c in range(1, min(sheet.max_column + 1, 30)):
            val = str(sheet.cell(row=r, column=c).value or '').strip()
            if any(kw in val for kw in ['编码', '代码', '名称', '品名']):
                return r
    return 1


def locate_col(sheet, header_row_idx):
    """识别列位置"""
    col_code = None
    col_name = None
    col_qty = None
    col_in_date = None
    col_prod_date = None
    col_kg = None

    for c in range(1, sheet.max_column + 1):
        val = str(sheet.cell(row=header_row_idx, column=c).value or '').strip().lower()
        if not val:
            continue
        if ('产品代码' in val or '产品编号' in val or '产品编码' in val
                or '商品编码' in val or '商品编号' in val or '商品代码' in val
                or '货品编码' in val or '货号' in val
                or ('编码' in val and '社会信用' not in val)
                or ('代码' in val and '社会信用' not in val)):
            col_code = c
        elif ('产品名称' in val or '商品名称' in val or '货品名称' in val
              or ('名称' in val and '单位' not in val and '编号' not in val)):
            col_name = c
        elif '采购数量' in val or '进货数量' in val or '箱数' in val or '件数' in val:
            col_qty = c
        elif '数量' in val and '单位' not in val:
            col_qty = c
        elif '进货日期' in val or '采购日期' in val or '送货日期' in val or '到货日期' in val:
            col_in_date = c
        elif '生产日期' in val:
            col_prod_date = c
        elif '千克' in val or '公斤' in val or 'kg' in val:
            col_kg = c

    return col_code, col_name, col_qty, col_in_date, col_prod_date, col_kg


def get_cell_cached_value(sheet_data_only, row, col):
    """从 data_only 模式的 sheet 读取缓存值"""
    val = sheet_data_only.cell(row=row, column=col).value
    if val is None:
        return ''
    return str(val).strip()


def main():
    if len(sys.argv) < 4:
        print("用法: python generate_ledger_bridge.py <template_path> <output_path> <rows_json_path>")
        sys.exit(1)

    template_path = sys.argv[1]
    output_path = sys.argv[2]
    rows_json_path = sys.argv[3]

    # 解析 rows JSON
    try:
        with open(rows_json_path, 'r', encoding='utf-8') as f:
            rows = json.load(f)
    except Exception as e:
        print(f"[Python] 读取 rows JSON 失败: {e}")
        sys.exit(1)

    print(f"[Python] 打开模板: {template_path}")
    print(f"[Python] 待填入商品数: {len(rows)}")

    # ★ 关键：用两种模式打开模板 ★
    # 1. data_only=True：读取公式的缓存值（用于匹配编码）
    wb_data = load_workbook(template_path, data_only=True)
    # 2. 普通模式：保留公式结构（用于写入数据）
    wb = load_workbook(template_path)
    print(f"[Python] 工作簿中的 sheet: {wb.sheetnames}")

    # 找到"进货表" sheet
    target_sheet = None
    target_sheet_data = None
    for sheet_name in wb.sheetnames:
        if '进货' in sheet_name:
            target_sheet = wb[sheet_name]
            target_sheet_data = wb_data[sheet_name]
            print(f"[Python] 找到进货表: {sheet_name}")
            break
    if target_sheet is None:
        target_sheet = wb.active
        target_sheet_data = wb_data.active
        print(f"[Python] 未找到进货表，使用活动 sheet: {target_sheet.title}")

    # 找表头行（用 data_only 版，因为标题行不应该有公式）
    header_row_idx = find_header_row(target_sheet, 15)
    print(f"[Python] 表头行: 第{header_row_idx}行")

    # 打印表头行所有值（诊断用）
    header_vals = []
    for c in range(1, min(target_sheet.max_column + 1, 20)):
        val = str(target_sheet.cell(row=header_row_idx, column=c).value or '').strip()
        header_vals.append(val)
    print(f"[Python] 表头值: {header_vals}")

    # 识别列位置
    col_code, col_name, col_qty, col_in_date, col_prod_date, col_kg = locate_col(
        target_sheet, header_row_idx
    )
    print(f"[Python] 列位置: code={col_code}, name={col_name}, qty={col_qty}, "
          f"in_date={col_in_date}, prod_date={col_prod_date}, kg={col_kg}")

    if col_code is None and col_name is None:
        print("[Python] 错误: 模板缺少商品编码列和名称列")
        sys.exit(1)

    # ★ 自动检测：编码列的值是公式还是实际值 ★
    # 如果编码列包含VLOOKUP公式，则需要从data_only版读取缓存值来匹配
    code_col_has_formula = False
    if col_code is not None:
        sample_val = target_sheet.cell(row=header_row_idx + 1, column=col_code).value
        if sample_val and str(sample_val).startswith('='):
            code_col_has_formula = True
            print(f"[Python] 编码列({col_code})包含公式，将从缓存值匹配")

    # 同时也检查A列（产品名称列）是否可以直接作为匹配键
    # 有些模板A列=名称，B列=编码(VLOOKUP)，需要用A列的名称匹配
    name_col_has_value = False
    if col_name is not None and col_code is not None:
        sample_name = get_cell_cached_value(target_sheet_data, header_row_idx + 1, col_name)
        if sample_name and not sample_name.startswith('='):
            name_col_has_value = True

    # 构建 rec_map
    rec_map = {}       # 精确编码 -> 行数据
    rec_map_norm = {}  # 规范化编码 -> 行数据
    rec_map_name = {}  # 产品名称 -> 行数据
    for row_data in rows:
        code = str(row_data.get('productCode', '')).strip()
        name = str(row_data.get('productName', '')).strip()
        if code:
            rec_map[code] = row_data
            nc = normalize_code(code)
            if nc:
                rec_map_norm[nc] = row_data
        if name:
            rec_map_name[name] = row_data

    print(f"[Python] rec_map 大小: {len(rec_map)}, rec_map_norm: {len(rec_map_norm)}, rec_map_name: {len(rec_map_name)}")

    # 日期
    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    today_str = today.isoformat()
    yesterday_str = yesterday.isoformat()

    # ★ 遍历模板现有行，按编码或名称匹配 ★
    data_start_row = header_row_idx + 1
    written_rows = set()
    updated_count = 0

    # 收集每行的匹配信息
    row_matches = {}  # row_idx -> (rec, code_raw)

    for r in range(data_start_row, target_sheet.max_row + 1):
        # 从 data_only 版读取缓存值（编码列可能是VLOOKUP公式）
        code_raw = get_cell_cached_value(target_sheet_data, r, col_code) if col_code else ''
        name_raw = get_cell_cached_value(target_sheet_data, r, col_name) if col_name else ''

        # 如果编码列为空但名称列有值，用名称匹配
        rec = None
        match_key = ''

        if code_raw and code_raw != 'None':
            code_norm = normalize_code(code_raw)
            # 先精确匹配，再规范化匹配
            rec = rec_map.get(code_raw) or rec_map_norm.get(code_norm)
            match_key = code_raw

        # 编码没匹配上，尝试名称匹配
        if rec is None and name_raw and name_raw != 'None':
            rec = rec_map_name.get(name_raw)
            if rec:
                match_key = name_raw
                print(f"[Python] 通过名称匹配: r={r}, name={name_raw}")

        if rec is None:
            # 未匹配，跳过（不删除行，保留模板结构）
            continue

        # ★ 匹配成功！写入数据到普通模式的workbook ★
        # 使用编码做72判断
        effective_code = str(rec.get('productCode', '')).strip()
        code_norm_for_check = normalize_code(effective_code)

        # ★ 重要：如果编码列是VLOOKUP公式，替换为实际值 ★
        # openpyxl 保存后会丢失公式缓存值，导致 Excel 打开时该列为空
        # 解决：直接写入实际代码值，不再保留 VLOOKUP 公式
        if col_code is not None and code_col_has_formula and effective_code:
            target_sheet.cell(row=r, column=col_code).value = effective_code

        # 同理：如果其他列也是VLOOKUP公式，也替换为缓存值
        # 例如：统一社会信用代码列、生产厂商列等
        for c in range(1, target_sheet.max_column + 1):
            cell = target_sheet.cell(row=r, column=c)
            if cell.value and str(cell.value).startswith('='):
                # 从data_only版读取缓存值
                cached = target_sheet_data.cell(row=r, column=c).value
                if cached is not None:
                    cell.value = cached

        # 数量
        if effective_code.startswith('72') or code_norm_for_check.startswith('72'):
            qty = 1
        else:
            qty = rec.get('boxCount', 0)

        if col_qty is not None:
            target_sheet.cell(row=r, column=col_qty).value = qty

        # 进货日期 = 今天
        if col_in_date is not None:
            target_sheet.cell(row=r, column=col_in_date).value = today_str

        # 生产日期
        if col_prod_date is not None:
            if effective_code.startswith('72') or code_norm_for_check.startswith('72'):
                target_sheet.cell(row=r, column=col_prod_date).value = yesterday_str
            else:
                prod_date = rec.get('productionDate', '')
                if prod_date:
                    try:
                        s = str(prod_date).replace('.', '-').replace('/', '-')
                        s = s[:10]
                        datetime.date.fromisoformat(s)
                        target_sheet.cell(row=r, column=col_prod_date).value = s
                    except Exception:
                        target_sheet.cell(row=r, column=col_prod_date).value = str(prod_date)

        written_rows.add(r)
        row_matches[r] = (rec, effective_code)
        updated_count += 1

    print(f"[Python] 模板行处理完成: 更新{updated_count}行")

    # ★ 删除未匹配的行 ★
    # 本次送货单未包含的商品行直接删除（未写入采购数量/进货日期/生产日期）
    rows_to_delete = []
    for r in range(data_start_row, target_sheet.max_row + 1):
        if r not in written_rows:
            rows_to_delete.append(r)

    # 从底部往上删（避免行号错乱）
    rows_to_delete.sort(reverse=True)
    for r in rows_to_delete:
        target_sheet.delete_rows(r, 1)

    print(f"[Python] 删除未匹配行: {len(rows_to_delete)}行")

    # ★ 追加送货单中模板不存在的商品 ★
    consumed_codes = set()
    consumed_codes_norm = set()
    consumed_names = set()
    for r, (rec, code) in row_matches.items():
        if code:
            consumed_codes.add(code)
            consumed_codes_norm.add(normalize_code(code))
        name = str(rec.get('productName', '')).strip()
        if name:
            consumed_names.add(name)

    appended = 0
    append_row = target_sheet.max_row + 1
    for rec in rows:
        code = str(rec.get('productCode', '')).strip()
        nc = normalize_code(code)
        name = str(rec.get('productName', '')).strip()

        if code in consumed_codes or nc in consumed_codes_norm or name in consumed_names:
            continue

        # 追加新行
        if col_name is not None:
            target_sheet.cell(row=append_row, column=col_name).value = name
        if col_code is not None:
            # 如果编码列是VLOOKUP公式列，直接写实际值而不是公式
            target_sheet.cell(row=append_row, column=col_code).value = code
        if col_qty is not None:
            target_sheet.cell(row=append_row, column=col_qty).value = rec.get('boxCount', 0)
        if col_in_date is not None:
            target_sheet.cell(row=append_row, column=col_in_date).value = rec.get('deliveryDate', today_str)
        if col_prod_date is not None:
            target_sheet.cell(row=append_row, column=col_prod_date).value = rec.get('productionDate', '')
        append_row += 1
        appended += 1

    if appended > 0:
        print(f"[Python] 追加模板中不存在的商品: {appended}条")

    # ★ 保存（openpyxl 保存时对命名范围很宽容，不会报错） ★
    print(f"[Python] 保存到: {output_path}")
    wb.save(output_path)
    print(f"[Python] 完成！共处理{updated_count}行匹配 + {appended}行追加")
    return 0


if __name__ == '__main__':
    try:
        sys.exit(main())
    except Exception as e:
        print(f"[Python] 错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
